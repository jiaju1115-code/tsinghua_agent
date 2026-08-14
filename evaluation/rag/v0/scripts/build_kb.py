from __future__ import annotations

import argparse
import hashlib
import json
import platform
import re
import sys
import time
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import joblib
import scipy
import sklearn
from openpyxl import load_workbook
from scipy import sparse
from sklearn.feature_extraction.text import TfidfVectorizer


RAG_DIR = Path(__file__).resolve().parents[1]
DATA_ROOT = RAG_DIR.parent
DEFAULT_CONFIG = RAG_DIR / "config" / "rag_v0.json"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def source_hash(text: str) -> str:
    """Match the existing manifests: trim and collapse every whitespace run."""
    canonical = re.sub(r"\s+", " ", text.strip())
    return sha256_bytes(canonical.encode("utf-8"))


def normalize_document(text: str) -> str:
    text = unicodedata.normalize("NFKC", text).replace("\r\n", "\n").replace("\r", "\n")
    lines: list[str] = []
    for raw in text.split("\n"):
        line = re.sub(r"[\t \u3000]+", " ", raw).strip()
        # Restricted captures may contain acquisition metadata. It is not useful
        # retrieval content and must never be propagated as credential material.
        if re.match(r"^-\s*(Auth|Discovery)\s*:", line, flags=re.I):
            continue
        lines.append(line)
    text = "\n".join(lines)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text


def read_sheet_rows(path: Path, key_header: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_index = next(
        i for i, row in enumerate(rows) if row and any(str(v).strip() == key_header for v in row if v is not None)
    )
    headers = [str(v).strip() if v is not None else "" for v in rows[header_index]]
    result: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not row or not any(v is not None and str(v).strip() for v in row):
            continue
        item = {h: v for h, v in zip(headers, row) if h}
        result.append(item)
    return result


def load_sources(config: dict[str, Any]) -> list[dict[str, Any]]:
    public_path = DATA_ROOT / config["public_manifest"]
    restricted_path = DATA_ROOT / config["restricted_manifest"]
    public = read_sheet_rows(public_path, "id")
    restricted = read_sheet_rows(restricted_path, "restricted_id")

    sources: list[dict[str, Any]] = []
    for row in public:
        if str(row.get("v3_2_action", "")).lower() != "approve":
            continue
        sources.append(
            {
                "source_id": str(row["id"]),
                "source_type": "public_staging",
                "title": str(row.get("title") or ""),
                "url": str(row.get("url") or ""),
                "category": str(row.get("category") or ""),
                "original_file": str(Path("staging_public_baseline_v1") / str(row["content_file"])).replace("\\", "/"),
                "declared_content_hash": str(row.get("content_hash") or ""),
                "v3_2_decision": str(row.get("v3_2_action") or ""),
            }
        )
    for row in restricted:
        if str(row.get("v3_2_action", "")).lower() != "approve":
            continue
        sources.append(
            {
                "source_id": str(row["restricted_id"]),
                "source_type": "restricted_approved",
                "title": str(row.get("title") or ""),
                "url": str(row.get("url") or ""),
                "category": str(row.get("category") or ""),
                "original_file": str(Path("restricted_expansion_v1") / str(row["source_file"])).replace("\\", "/"),
                "declared_content_hash": str(row.get("content_hash") or ""),
                "v3_2_decision": str(row.get("v3_2_action") or ""),
            }
        )
    return sorted(sources, key=lambda x: x["source_id"])


def pick_break(text: str, start: int, hard_end: int, min_end: int) -> int:
    window = text[start:hard_end]
    candidates: list[int] = []
    for match in re.finditer(r"\n\n|[。！？!?；;]\s*|\n", window):
        pos = start + match.end()
        if pos >= min_end:
            candidates.append(pos)
    return candidates[-1] if candidates else hard_end


def chunk_text(text: str, max_chars: int, overlap_chars: int, min_chars: int) -> list[tuple[int, int, str]]:
    if not text:
        return []
    chunks: list[tuple[int, int, str]] = []
    start = 0
    n = len(text)
    while start < n:
        hard_end = min(n, start + max_chars)
        if hard_end < n:
            end = pick_break(text, start, hard_end, start + int(max_chars * 0.55))
        else:
            end = n
        piece = text[start:end].strip()
        if piece:
            chunks.append((start, end, piece))
        if end >= n:
            break
        next_start = max(start + 1, end - overlap_chars)
        while next_start < end and not text[next_start].isspace():
            next_start += 1
        while next_start < n and text[next_start].isspace():
            next_start += 1
        start = next_start
    if len(chunks) > 1 and len(chunks[-1][2]) < min_chars:
        s0, _, p0 = chunks[-2]
        _, e1, p1 = chunks[-1]
        merged = (p0 + "\n\n" + p1).strip()
        if len(merged) <= max_chars + min_chars:
            chunks[-2:] = [(s0, e1, merged)]
    return chunks


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=False) + "\n")


def embed_texts(
    texts: list[str], embedding: dict[str, Any]
) -> tuple[sparse.csr_matrix, TfidfVectorizer, dict[str, Any]]:
    started = time.perf_counter()
    vectorizer = TfidfVectorizer(
        analyzer=str(embedding["analyzer"]),
        ngram_range=tuple(int(x) for x in embedding["ngram_range"]),
        max_features=int(embedding["max_features"]),
        sublinear_tf=bool(embedding["sublinear_tf"]),
        norm=str(embedding["norm"]),
        dtype=np.float32,
        lowercase=True,
    )
    matrix = vectorizer.fit_transform(texts).tocsr().astype(np.float32)
    metadata = {
        "backend": embedding["backend"],
        "analyzer": embedding["analyzer"],
        "ngram_range": embedding["ngram_range"],
        "max_features": embedding["max_features"],
        "normalization": embedding["norm"],
        "device": "cpu",
        "dimension": int(matrix.shape[1]),
        "row_count": int(matrix.shape[0]),
        "nonzero_values": int(matrix.nnz),
        "build_seconds": round(time.perf_counter() - started, 3),
    }
    return matrix, vectorizer, metadata


def build(config_path: Path) -> dict[str, Any]:
    config = json.loads(config_path.read_text(encoding="utf-8"))
    sources = load_sources(config)
    validation_rows: list[dict[str, Any]] = []
    accepted: list[dict[str, Any]] = []
    normalized_documents: dict[str, str] = {}

    for source in sources:
        file_path = DATA_ROOT / source["original_file"]
        exists = file_path.is_file()
        raw = file_path.read_bytes() if exists else b""
        try:
            text = raw.decode("utf-8") if raw else ""
            decode_ok = True
        except UnicodeDecodeError:
            text = ""
            decode_ok = False
        normalized = normalize_document(text) if decode_ok else ""
        computed_hash = source_hash(text) if text else ""
        hash_match = bool(computed_hash and computed_hash == source["declared_content_hash"])
        nonempty = bool(normalized.strip())
        status = "accepted" if exists and decode_ok and hash_match and nonempty else "rejected_validation"
        row = {
            **source,
            "file_exists": exists,
            "utf8_decode_ok": decode_ok,
            "content_nonempty": nonempty,
            "declared_hash_algorithm": "sha256(trim+collapse_all_whitespace)",
            "computed_content_hash": computed_hash,
            "hash_match": hash_match,
            "raw_file_sha256": sha256_bytes(raw) if raw else "",
            "normalized_chars": len(normalized),
            "normalized_document_sha256": sha256_bytes(normalized.encode("utf-8")) if normalized else "",
            "normalized_text_path": f"normalized_documents/{source['source_id']}.txt" if normalized else "",
            "validation_status": status,
        }
        validation_rows.append(row)
        if status == "accepted":
            accepted.append(row)
            normalized_documents[source["source_id"]] = normalized

    normalized_dir = RAG_DIR / "normalized_documents"
    normalized_dir.mkdir(parents=True, exist_ok=True)
    for source in accepted:
        (normalized_dir / f"{source['source_id']}.txt").write_text(
            normalized_documents[source["source_id"]], encoding="utf-8", newline="\n"
        )

    chunks: list[dict[str, Any]] = []
    chunk_cfg = config["chunking"]
    for source in accepted:
        parts = chunk_text(
            normalized_documents[source["source_id"]],
            int(chunk_cfg["max_chars"]),
            int(chunk_cfg["overlap_chars"]),
            int(chunk_cfg["min_chars"]),
        )
        for index, (char_start, char_end, text) in enumerate(parts):
            chunks.append(
                {
                    "chunk_id": f"PROV0-{source['source_id']}-{index:04d}",
                    "source_id": source["source_id"],
                    "title": source["title"],
                    "url": source["url"],
                    "category": source["category"],
                    "source_type": source["source_type"],
                    "original_file": source["original_file"],
                    "text": text,
                    "chunk_index": index,
                    "char_start": char_start,
                    "char_end": char_end,
                    "chunk_sha256": sha256_bytes(text.encode("utf-8")),
                }
            )

    write_jsonl(RAG_DIR / "knowledge_base_manifest" / "knowledge_base_manifest.jsonl", validation_rows)
    write_jsonl(RAG_DIR / "chunks" / "chunks.jsonl", chunks)

    embedding_inputs = [f"{row['title']}\n{row['text']}" for row in chunks]
    matrix, vectorizer, index_metadata = embed_texts(embedding_inputs, config["embedding"])
    index_dir = RAG_DIR / "vector_index"
    index_dir.mkdir(parents=True, exist_ok=True)
    sparse.save_npz(index_dir / "embeddings.npz", matrix, compressed=True)
    joblib.dump(vectorizer, index_dir / "tfidf_vectorizer.joblib", compress=3)
    (index_dir / "chunk_ids.json").write_text(
        json.dumps([row["chunk_id"] for row in chunks], ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    index_metadata.update(
        {
            "kb_name": config["kb_name"],
            "index_type": "sparse_flat_cosine_scipy_csr",
            "embedding_input": "title + newline + chunk_text",
            "embeddings_sha256": sha256_bytes((index_dir / "embeddings.npz").read_bytes()),
            "vectorizer_sha256": sha256_bytes((index_dir / "tfidf_vectorizer.joblib").read_bytes()),
            "chunks_sha256": sha256_bytes((RAG_DIR / "chunks" / "chunks.jsonl").read_bytes()),
            "created_at_utc": datetime.now(timezone.utc).isoformat(),
            "python": sys.version.split()[0],
            "platform": platform.platform(),
            "numpy": np.__version__,
            "scipy": scipy.__version__,
            "scikit_learn": sklearn.__version__,
        }
    )
    (index_dir / "index_metadata.json").write_text(
        json.dumps(index_metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    summary = {
        "kb_name": config["kb_name"],
        "manifest_approved_rows": len(sources),
        "accepted_sources": len(accepted),
        "rejected_sources": len(sources) - len(accepted),
        "public_accepted": sum(r["source_type"] == "public_staging" and r["validation_status"] == "accepted" for r in validation_rows),
        "restricted_accepted": sum(r["source_type"] == "restricted_approved" and r["validation_status"] == "accepted" for r in validation_rows),
        "chunk_count": len(chunks),
        "index": index_metadata,
    }
    (RAG_DIR / "knowledge_base_manifest" / "build_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return summary


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build the isolated PROVISIONAL_KB_V0 corpus and vector index.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    args = parser.parse_args()
    print(json.dumps(build(args.config.resolve()), ensure_ascii=False, indent=2))
