"""Join and summarize future human review inputs without assigning labels.

The script is read-only with respect to all inputs.  In particular, it never
derives or writes retrieval failure_type values for E2E-50 cases.
"""
from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any


FAILURE_TYPES = {"kb_missing", "retrieval_miss", "ranking_miss", "entity_mismatch", "evidence_over_refusal", "other"}
REFUSAL_TYPES = {"CORRECT_REFUSAL": "correct_refusal", "OVER_REFUSAL": "over_refusal", "SHOULD_REFUSE_BUT_ANSWERED": "should_refuse_but_answered"}


def load_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".jsonl":
        return [json.loads(line) for line in path.read_text(encoding="utf-8-sig").splitlines() if line.strip()]
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SystemExit(".xlsx support requires openpyxl; use a JSONL export if unavailable") from exc
        sheet = load_workbook(path, read_only=True, data_only=True).active
        values = list(sheet.iter_rows(values_only=True))
        return [dict(zip(values[0], row)) for row in values[1:] if any(cell is not None for cell in row)]
    raise SystemExit(f"unsupported input format: {path}")


def index(rows: list[dict[str, Any]], label: str) -> dict[str, dict[str, Any]]:
    result = {}
    for row in rows:
        case_id = row.get("case_id")
        if not isinstance(case_id, str) or not case_id:
            raise ValueError(f"{label} row missing case_id")
        if case_id in result:
            raise ValueError(f"duplicate {label} case_id: {case_id}")
        result[case_id] = row
    return result


def confusion(rows: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
    matrix: dict[str, Counter] = {}
    for row in rows:
        human = row.get("human_evidence_sufficiency")
        runtime = row["result"].get("evidence_status")
        if human is not None and runtime is not None:
            matrix.setdefault(str(human), Counter())[str(runtime)] += 1
    return {key: dict(value) for key, value in sorted(matrix.items())}


def summarize(review: list[dict[str, Any]], template: list[dict[str, Any]], replay: list[dict[str, Any]], results: list[dict[str, Any]]) -> dict[str, Any]:
    review_by_case, template_by_case, replay_by_case, results_by_case = (index(rows, name) for rows, name in ((review, "review"), (template, "template"), (replay, "replay"), (results, "results")))
    joined_ids = sorted(set(review_by_case) & set(template_by_case) & set(replay_by_case) & set(results_by_case))
    joined = [{"review": review_by_case[case_id], "template": template_by_case[case_id], "replay": replay_by_case[case_id], "result": results_by_case[case_id]} for case_id in joined_ids]
    invalid_failure_values = sorted({row["template"].get("failure_type") for row in joined if row["template"].get("failure_type") is not None and row["template"].get("failure_type") not in FAILURE_TYPES})
    if invalid_failure_values:
        raise ValueError(f"invalid human failure_type values: {invalid_failure_values}")
    refusal = Counter(REFUSAL_TYPES.get(row["review"].get("refusal_appropriateness")) for row in joined)
    failure = Counter(row["template"].get("failure_type") for row in joined)
    downstream = Counter()
    for row in joined:
        if row["template"].get("failure_type") == "evidence_over_refusal": downstream["evidence_over_refusal"] += 1
        if row["review"].get("citation_correctness") == 0: downstream["citation_error"] += 1
        if row["review"].get("answer_correctness") == 0: downstream["answer_error"] += 1
        if row["template"].get("failure_type") == "other": downstream["other"] += 1
    return {
        "status": "HUMAN_REVIEW_JOIN_READY_NO_AUTO_ATTRIBUTION",
        "input_counts": {"review": len(review), "template": len(template), "replay": len(replay), "results": len(results), "joined": len(joined)},
        "unjoined_case_counts": {"review": len(set(review_by_case) - set(joined_ids)), "template": len(set(template_by_case) - set(joined_ids)), "replay": len(set(replay_by_case) - set(joined_ids)), "results": len(set(results_by_case) - set(joined_ids))},
        "evidence_comparison": {"human_evidence_vs_evidence_v1_confusion_matrix": confusion(joined)},
        "refusal": {key: refusal.get(key, 0) for key in ("correct_refusal", "over_refusal", "should_refuse_but_answered")},
        "retrieval_failure": {key: failure.get(key, 0) for key in ("kb_missing", "retrieval_miss", "ranking_miss", "entity_mismatch")},
        "downstream": {key: downstream.get(key, 0) for key in ("evidence_over_refusal", "citation_error", "answer_error", "other")},
        "human_failure_type_missing": failure.get(None, 0),
        "auto_attribution_performed": False,
    }


def fixture() -> dict[str, Any]:
    review = [{"case_id": "FIX-001", "refusal_appropriateness": "OVER_REFUSAL", "citation_correctness": 0, "answer_correctness": 1, "human_evidence_sufficiency": "SUFFICIENT"}]
    template = [{"case_id": "FIX-001", "failure_type": None}]
    replay = [{"case_id": "FIX-001"}]
    results = [{"case_id": "FIX-001", "evidence_status": "INSUFFICIENT"}]
    return summarize(review, template, replay, results)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--review", type=Path)
    parser.add_argument("--template", type=Path)
    parser.add_argument("--replay", type=Path)
    parser.add_argument("--results", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--fixture", action="store_true")
    args = parser.parse_args()
    if args.fixture:
        summary = fixture()
    else:
        if not all((args.review, args.template, args.replay, args.results)):
            raise SystemExit("--review --template --replay and --results are required unless --fixture is used")
        summary = summarize(load_rows(args.review), load_rows(args.template), load_rows(args.replay), load_rows(args.results))
    if args.output:
        if args.output.exists():
            raise SystemExit(f"refusing to overwrite output: {args.output}")
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
