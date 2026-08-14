from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(r"D:\python_projects\tsinghua_ai")
SECOND = ROOT / "data_second"
TEST = SECOND / "prompt_v3_2_test"
HUMAN_XLSX = SECOND / "public_rebuild_v1" / "human_check" / "public_rebuild_v1_human_check.xlsx"

TOPIC_IRRELEVANT = {
    "PUBFOLLOW000007": "页面核心是一次具体展览，属于低复用文化活动；即使暂不判断日期，也不是目标知识库的长期条目。",
    "PUBEXP000186": "主体为其他高校的开放获取政策，不是清华自身校园知识。",
    "PUBEXP000187": "主体为其他高校的开放获取政策，不是清华自身校园知识。",
    "PUBEXP000184": "主体是泛开放科学评论，不能回答清华校园服务、资源、机构或运行问题。",
    "PUBEXP000051": "主体是外部开放获取组织观点解读，和清华只有发布站点关系。",
    "PUBEXP000201": "主体是外部开放获取议题评论，缺少清华校园知识。",
    "PUBEXP000119": "页面核心是一次 IEEE 专场活动报道，不是持续存在的图书馆服务或科研资源页。",
    "PUBEXP000127": "页面核心是一期真人图书馆活动及人物交流，属于低复用活动报道。",
    "PUBEXP000128": "页面核心是一期真人图书馆嘉宾活动，属于低复用人物/活动报道。",
    "PUBEXP000293": "页面核心是一期嘉宾经历分享，不能帮助用户理解或使用稳定校园资源。",
    "PUBEXP000299": "页面核心是一期嘉宾活动及人物内容，属于低复用活动报道。",
    "PUBEXP000188": "主体为其他高校的开放获取政策，不是清华自身校园知识。",
    "PUBEXP000190": "主体为其他高校的开放获取政策，不是清华自身校园知识。",
    "PUBEXP000233": "页面核心是单场经济金融数据课堂，而不是数据库正式资源页或长期使用指南。",
    "PUBEXP000158": "正文只提供专题/公告聚合性内容，没有可复用的具体校园知识条目。",
}

EXPIRED_RELEVANT = {
    "PUBFOLLOW000003": "图书馆荐购机会本身与学生使用校园资源有关，但该单次现场活动和报名窗口已结束。",
    "PUBFOLLOW000002": "领奖安排属于校园用户事务，但领取窗口已经结束，当前已无使用价值。",
    "PUBEXP000070": "图书馆论文写作培训属于学生学习支持，主题高度相关，但本学期培训安排已经结束。",
}

MEDIUM = {
    "PUBEXP000221": "图书馆专题书架属于校园文化与图书馆服务，具有一定复用价值，但持续期限和长期性需结合时效复核。",
}


def load_human():
    ws = openpyxl.load_workbook(HUMAN_XLSX, data_only=True).active
    headers = [c.value for c in ws[3]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=4, values_only=True) if row[0]]


def main():
    rows = []
    for h in load_human():
        ident = str(h["id"])
        human_action = str(h.get("human_action") or "").lower()
        if ident in TOPIC_IRRELEVANT:
            relevance, reject_type, reason = "low", "topic_irrelevant", TOPIC_IRRELEVANT[ident]
        elif ident in EXPIRED_RELEVANT:
            relevance, reject_type, reason = "high", "expired_event", EXPIRED_RELEVANT[ident]
        elif ident in MEDIUM:
            relevance, reject_type, reason = "medium", "", MEDIUM[ident]
        else:
            relevance, reject_type = "high", ""
            reason = "页面核心是清华校园服务、机构、规则、数据库、信息系统或学生可利用的资源，属于目标知识库范围。"
        question = ""
        if ident == "PUBEXP000158":
            question = "人工备注以‘只列公告、没有具体正文’为由拒绝，部分涉及 Quality Gate 职责；应复核该页为何进入已通过正文质量门槛的样本。"
        rows.append({
            "id": ident,
            "title": h.get("title") or "",
            "url": h.get("url") or "",
            "human_action": human_action,
            "human_note": h.get("human_note") or "",
            "proposed_topic_relevance": relevance,
            "proposed_reject_type": reject_type,
            "analysis_reason": reason,
            "human_label_question": question,
        })
    if len(rows) != 30:
        raise RuntimeError(f"expected 30 samples, got {len(rows)}")
    summary = {
        "samples": len(rows),
        "human_actions": dict(Counter(x["human_action"] for x in rows)),
        "proposed_topic_relevance": dict(Counter(x["proposed_topic_relevance"] for x in rows)),
        "human_reject_proposed_types": dict(Counter(x["proposed_reject_type"] for x in rows if x["human_action"] == "reject")),
        "human_label_questions": sum(bool(x["human_label_question"]) for x in rows),
    }
    (TEST / "analysis").mkdir(parents=True, exist_ok=True)
    (TEST / "audit").mkdir(parents=True, exist_ok=True)
    (TEST / "analysis" / "v3_2_topic_analysis.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    (TEST / "audit" / "topic_analysis_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
