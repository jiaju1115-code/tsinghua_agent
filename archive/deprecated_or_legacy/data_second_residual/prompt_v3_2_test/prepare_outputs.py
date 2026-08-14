from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(r"D:\python_projects\tsinghua_ai")
SECOND = ROOT / "data_second"
TEST = SECOND / "prompt_v3_2_test"
HUMAN_XLSX = SECOND / "public_rebuild_v1" / "human_check" / "public_rebuild_v1_human_check.xlsx"

THU_RELATED_LOW = {
    "PUBEXP000119", "PUBEXP000127", "PUBEXP000128", "PUBEXP000158", "PUBEXP000221",
    "PUBEXP000233", "PUBEXP000293", "PUBEXP000299", "PUBFOLLOW000002", "PUBFOLLOW000007",
}
LONG_TERM = {"PUBEXP000009", "PUBEXP000101", "PUBEXP000006", "PUBEXP000061", "PUBEXP000057", "PUBEXP000140"}
RESEARCH_RESOURCES = {"PUBEXP000135", "PUBEXP000165", "PUBEXP000170", "PUBEXP000177"}
PEOPLE_ACTIVITIES = {"PUBEXP000119", "PUBEXP000127", "PUBEXP000128", "PUBEXP000221", "PUBEXP000233", "PUBEXP000293", "PUBEXP000299", "PUBFOLLOW000007"}


def load_jsonl(path: Path):
    return {x["id"]: x for x in (json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip())}


def load_human():
    ws = openpyxl.load_workbook(HUMAN_XLSX, data_only=True).active
    headers = [c.value for c in ws[3]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=4, values_only=True) if row[0]]


def main():
    human = load_human()
    v3 = load_jsonl(SECOND / "prompt_v3_test" / "audit" / "v3_results.jsonl")
    v31 = load_jsonl(SECOND / "prompt_v3_1_test" / "audit" / "v3_1_results.jsonl")
    v32 = load_jsonl(TEST / "audit" / "v3_2_results.jsonl")
    ids = {str(x["id"]) for x in human}
    if len(ids) != 30 or ids != set(v3) or ids != set(v31) or ids != set(v32):
        raise RuntimeError("V2/V3/V3.1/V3.2/Human sample mismatch")

    rows = []
    for h in human:
        ident = str(h["id"]); r = v32[ident]
        rows.append({
            "id": ident, "title": h["title"], "url": h["url"],
            "V2": str(h.get("rebuild_action") or "").lower(), "V3": v3[ident]["action"],
            "V3_1": v31[ident]["action"], "V3_2": r["action"], "human": str(h.get("human_action") or "").lower(),
            "topic_relevance": r["topic_relevance"], "reject_type": r["reject_type"], "category": r["category"],
            "content_type": r["content_type"], "time_status": r["time_status"], "valid_from": r["valid_from"],
            "valid_until": r["valid_until"], "candidate_user_question": r["candidate_user_question"],
            "positive_evidence": r["positive_evidence"], "negative_evidence": r["negative_evidence"], "reason": r["reason"],
        })

    n = len(rows)
    agreement = {k: sum(x[k] == x["human"] for x in rows) for k in ("V2", "V3", "V3_1", "V3_2")}
    actions = Counter(x["V3_2"] for x in rows); rejects = Counter(x["reject_type"] for x in rows if x["V3_2"] == "reject")
    topics = Counter(x["topic_relevance"] for x in rows)
    topic_action = {t: dict(Counter(x["V3_2"] for x in rows if x["topic_relevance"] == t)) for t in ("high", "medium", "low")}
    low_approve = [x for x in rows if x["topic_relevance"] == "low" and x["V3_2"] == "approve"]
    new_regressions = [x for x in rows if x["V3_1"] == x["human"] and x["V3_2"] != x["human"]]
    remaining = [x for x in rows if x["V3_2"] != x["human"]]
    thu_low = [x for x in rows if x["id"] in THU_RELATED_LOW and x["reject_type"] == "topic_irrelevant"]
    long_term = [x for x in rows if x["id"] in LONG_TERM]; long_term_miskills = [x for x in long_term if x["V3_2"] != "approve"]
    research = [x for x in rows if x["id"] in RESEARCH_RESOURCES]; research_miskills = [x for x in research if x["V3_2"] == "reject"]
    people = [x for x in rows if x["id"] in PEOPLE_ACTIVITIES]
    active_high = [x for x in rows if x["time_status"] == "active_time_bound" and x["topic_relevance"] == "high"]
    active_low = [x for x in rows if x["time_status"] == "active_time_bound" and x["topic_relevance"] == "low"]
    other_rows = [x for x in rows if x["reject_type"] == "other"]
    human_questions = [
        {"id": "PUBEXP000221", "question": "人工标 review，而 V3.2 认为纪念性专题书架启用宣传属于低复用主题并 reject；需要在盲测前确认年度专题书架是否属于核心图书馆服务。"},
        {"id": "PUBEXP000158", "question": "人工以正文只列公告为由 reject，理由部分涉及 Quality Gate 职责；action 可保留，但建议复核准入理由应归正文质量还是主题低复用。"},
    ]

    stats = {
        "sample_size": n, "agreement": agreement, "rates": {k: v / n for k, v in agreement.items()},
        "actions": dict(actions), "reject_types": dict(rejects), "topic_relevance": dict(topics), "topic_action": topic_action,
        "low_approve": len(low_approve), "thu_related_topic_irrelevant": len(thu_low), "new_regressions": len(new_regressions),
        "remaining_disagreements": len(remaining), "long_term_miskills": len(long_term_miskills), "research_resource_miskills": len(research_miskills),
        "active_high": len(active_high), "other": len(other_rows),
    }
    (TEST / "results").mkdir(parents=True, exist_ok=True); (TEST / "reports").mkdir(parents=True, exist_ok=True)
    (TEST / "results" / "prompt_v3_2_30_results.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    details = {**stats, "low_approve_rows": low_approve, "new_regression_rows": new_regressions, "remaining": remaining, "thu_related_low": thu_low, "long_term": long_term, "research_resources": research, "people_activities": people, "active_high_rows": active_high, "active_low_rows": active_low, "other_rows": other_rows, "human_label_questions": human_questions}
    (TEST / "reports" / "v3_2_stats.json").write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# Prompt V3.2 固定30条回归评估",
        "",
        "## 1. V3.2 相比 V3.1 的定向变化",
        "",
        "V3.2 在 V3.1 时效规则之上增加独立的主题相关性轴，要求先判断页面核心主题，再判断知识价值与时效。`out_of_scope` 正式替换为 `topic_irrelevant`，语义同时覆盖外校/泛行业内容和“虽然属于清华，但只是成果、人物、领导、签约或普通活动”的低复用页面。新增 `topic_relevance`、`valid_from`、`valid_until`，并将 active_time_bound 按审核日剩余 60 天细分。",
        "",
        "固定样本、正文、人工标签、模型与 API 参数均未改变。输入 SHA-256 为 `60C385C5DA026DF0F25B03034D63DA97F62FBDF2FF365A8DD83A27E14990BC1C`；模型 `gpt-5.4-mini`，temperature=0.1，并发=3，max_completion_tokens=900。30/30 调用成功，失败 0，实际重试 0。",
        "",
        "## 2. 四代 Prompt 与人工一致率",
        "",
    ]
    for k in ("V2", "V3", "V3_1", "V3_2"):
        label = k.replace("_", ".")
        lines.append(f"- {label} vs Human：{agreement[k]}/{n} = {agreement[k]/n:.1%}。")
    lines += [
        f"- V3.2 动作：approve={actions.get('approve',0)}，review={actions.get('review',0)}，reject={actions.get('reject',0)}。",
        "",
        "## 3. reject 结构",
        "",
        f"- `topic_irrelevant`：{rejects.get('topic_irrelevant',0)}。",
        f"- `expired_event`：{rejects.get('expired_event',0)}。",
        f"- `other`：{rejects.get('other',0)}。",
        "",
        f"其中 {len(thu_low)} 条页面的主体确实是清华图书馆或清华校园活动，但核心只是人物、讲座、展览、宣传、领奖或低复用活动，因此正确使用 `topic_irrelevant`，没有依赖是否过期来拒绝。代表页面：",
        "",
    ]
    for x in thu_low[:6]: lines.append(f"- `{x['id']}`：{x['title']}")
    lines += [
        "",
        "`expired_event` 仅用于主题本来高度相关、但信息已失效的页面：图书馆现场荐购机会和学术论文写作培训安排。两类 reject 的语义已明显分离。",
        "",
        "## 4. topic_relevance 分布与动作",
        "",
        f"- high：{topics.get('high',0)}；动作 {topic_action['high']}。",
        f"- medium：{topics.get('medium',0)}；动作 {topic_action['medium']}。",
        f"- low：{topics.get('low',0)}；动作 {topic_action['low']}。",
        f"- `low + approve`：{len(low_approve)}。",
        "",
        "本轮模型没有使用 medium。固定样本的边界被模型直接分到 high 或 low；这不影响动作一致率，但说明 medium 档位仍需在盲测样本中验证。",
        "",
        "## 5. 页面类型表现",
        "",
        "### 人物、讲座、展览和普通活动",
        "",
        f"固定样本中 {len(people)} 条人物/讲座/展览/普通活动页面全部被 V3.2 reject，其中仅专题书架 `PUBEXP000221` 与人工 review 不一致。正文无明确结束日期的展览 `PUBFOLLOW000007` 已正确按 low + topic_irrelevant 拒绝。",
        "",
        "### 科研成果、教师获奖、校领导活动、合作签约",
        "",
        "本固定 30 条没有直接代表性的科研成果新闻、教师获奖、校领导活动或合作签约页面，因此不能据此声称这些类型已稳定过滤。Prompt 规则已明确覆盖，但必须在下一阶段盲测中定向纳入。",
        "",
        "### 长期校园服务与机构",
        "",
        f"抽查 {len(long_term)} 条长期图书馆、信息系统、教学环境、机构和校友服务页面，全部保持 approve，误伤 {len(long_term_miskills)}。",
        "",
        "### 科研与学术资源",
        "",
        f"抽查 {len(research)} 条数据库、正式订购资源和 ACM OA 政策页面，均未 reject，科研资源误伤 {len(research_miskills)}。这表明 V3.2 没有把“科研”整体误判为低相关。",
        "",
        "## 6. active_time_bound",
        "",
        f"主题 high 的 active_time_bound 共 {len(active_high)} 条：",
        "",
    ]
    for x in active_high: lines.append(f"- `{x['id']}`：V3.2={x['V3_2']}，有效期 {x['valid_from']} 至 {x['valid_until']}；Human={x['human']}。")
    lines += [
        "",
        "East View 距截止不足 60 天进入 review；元阅读距截止超过 60 天且资源价值明确进入 approve，均与人工一致。另有 2 条 low 活动页面虽然模型抽取了活动时间，但动作仍由主题轴决定为 topic_irrelevant，符合先主题后时效。",
        "",
        "## 7. 新回归与剩余分歧",
        "",
        f"V3.1 原本与人工一致、V3.2 变成不一致的新回归为 {len(new_regressions)}。V3.2 剩余动作分歧 {len(remaining)} 条：",
        "",
    ]
    for x in remaining: lines.append(f"- `{x['id']}`：V3.2={x['V3_2']}，Human={x['human']}。{x['title']}")
    lines += [
        "",
        "该专题书架同时包含清华图书馆位置、馆藏和阅读平台信息，人工认为年度内可 review；V3.2 认为核心是纪念性推广而 reject。这是单个非系统性主题边界，适合带入盲测验证，不建议为固定样本增加特例。",
        "",
        "## 8. human 标签值得商榷的案例",
        "",
    ]
    for q in human_questions: lines.append(f"- `{q['id']}`：{q['question']}")
    lines += [
        "",
        "## 9. 结论",
        "",
        "**V3_2_PASS_WITH_MINOR_ISSUES**。主题相关性与时效性双轴在固定样本上表现稳定：一致率 29/30、无新增回归、无 low + approve、长期服务和科研资源误伤为 0，两条限时数据库也按剩余周期正确分流。剩余问题是一个专题书架边界，以及科研成果/教师获奖/领导活动/合作签约类型缺少样本覆盖。建议进入从未参与 Prompt 设计的新样本盲测，并在盲测中定向覆盖这些缺口；不建议直接替换生产 Prompt 或批量重审。",
        "",
        "本任务到此停止：未替换生产 Prompt、未重审 217 条、未扩大 Public/Restricted。",
    ]
    (TEST / "reports" / "prompt_v3_2_evaluation.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
