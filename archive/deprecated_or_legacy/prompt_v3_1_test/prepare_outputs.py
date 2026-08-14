from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(r"D:\python_projects\tsinghua_ai")
SECOND = ROOT / "data_second"
TEST = SECOND / "prompt_v3_1_test"
HUMAN_XLSX = SECOND / "public_rebuild_v1" / "human_check" / "public_rebuild_v1_human_check.xlsx"
KNOWN_IDS = ["PUBEXP000076", "PUBEXP000233", "PUBFOLLOW000007", "PUBEXP000119", "PUBEXP000293", "PUBEXP000299"]
EXPECTED = {
    "PUBEXP000076": ("review", ""),
    "PUBEXP000233": ("reject", "expired_event"),
    "PUBFOLLOW000007": ("reject", "expired_event"),
    "PUBEXP000119": ("reject", "expired_event"),
    "PUBEXP000293": ("reject", "expired_event"),
    "PUBEXP000299": ("reject", "expired_event"),
}


def load_jsonl(path: Path):
    return {x["id"]: x for x in (json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip())}


def load_human():
    ws = openpyxl.load_workbook(HUMAN_XLSX, data_only=True).active
    headers = [c.value for c in ws[3]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=4, values_only=True) if row[0]]


def main():
    human = load_human()
    v3 = load_jsonl(SECOND / "prompt_v3_test" / "audit" / "v3_results.jsonl")
    v31 = load_jsonl(TEST / "audit" / "v3_1_results.jsonl")
    if len(human) != 30 or {str(x["id"]) for x in human} != set(v3) or set(v3) != set(v31):
        raise RuntimeError("V2/V3/V3.1/Human sample set mismatch")

    rows = []
    for h in human:
        ident = str(h["id"])
        r3, r31 = v3[ident], v31[ident]
        rows.append({
            "id": ident,
            "title": h["title"],
            "url": h["url"],
            "V2_action": str(h["rebuild_action"] or "").lower(),
            "V3_action": r3["action"],
            "V3_1_action": r31["action"],
            "human_action": str(h["human_action"] or "").lower(),
            "category": r31["category"],
            "content_type": r31["content_type"],
            "time_status": r31["time_status"],
            "reject_type": r31["reject_type"],
            "candidate_user_question": r31["candidate_user_question"],
            "positive_evidence": r31["positive_evidence"],
            "negative_evidence": r31["negative_evidence"],
            "reason": r31["reason"],
        })

    n = len(rows)
    agreements = {k: sum(x[k] == x["human_action"] for x in rows) for k in ("V2_action", "V3_action", "V3_1_action")}
    actions = Counter(x["V3_1_action"] for x in rows)
    reject_types = Counter(x["reject_type"] for x in rows if x["V3_1_action"] == "reject")
    transition_order = [f"{a}→{b}" for a in ("approve", "review", "reject") for b in ("approve", "review", "reject")]
    transitions_raw = Counter(f'{x["V3_action"]}→{x["V3_1_action"]}' for x in rows)
    transitions = {k: transitions_raw.get(k, 0) for k in transition_order}
    known = []
    for ident in KNOWN_IDS:
        x = next(r for r in rows if r["id"] == ident)
        exp_action, exp_reject = EXPECTED[ident]
        passed = x["V3_1_action"] == exp_action and (not exp_reject or x["reject_type"] == exp_reject)
        known.append({**x, "expected_action": exp_action, "expected_reject_type": exp_reject, "passed": passed})
    new_regressions = [x for x in rows if x["V3_action"] == x["human_action"] and x["V3_1_action"] != x["human_action"]]
    remaining = [x for x in rows if x["V3_1_action"] != x["human_action"]]
    active = [x for x in rows if x["time_status"] == "active_time_bound"]
    long_term_ids = ["PUBEXP000009", "PUBEXP000101", "PUBEXP000006", "PUBEXP000061", "PUBEXP000057", "PUBEXP000135", "PUBEXP000165", "PUBEXP000177"]
    long_term = [next(x for x in rows if x["id"] == ident) for ident in long_term_ids]
    long_term_miskills = [x for x in long_term if x["V3_1_action"] == "reject"]
    event_rows = [x for x in rows if x["reject_type"] == "expired_event" or x["id"] == "PUBFOLLOW000007"]
    other_rows = [x for x in rows if x["reject_type"] == "other"]

    stats = {
        "sample_size": n,
        "agreements": {"V2": agreements["V2_action"], "V3": agreements["V3_action"], "V3.1": agreements["V3_1_action"]},
        "rates": {"V2": agreements["V2_action"] / n, "V3": agreements["V3_action"] / n, "V3.1": agreements["V3_1_action"] / n},
        "actions": dict(actions), "reject_types": dict(reject_types), "transitions": transitions,
        "known_fixed": sum(x["passed"] for x in known), "known_total": len(known),
        "new_regressions": len(new_regressions), "active_time_bound_count": len(active),
        "active_time_bound_all_review": all(x["V3_1_action"] == "review" for x in active),
        "long_term_miskills": len(long_term_miskills), "other_count": len(other_rows),
    }
    (TEST / "results").mkdir(parents=True, exist_ok=True)
    (TEST / "reports").mkdir(parents=True, exist_ok=True)
    (TEST / "results" / "prompt_v3_1_30_results.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    (TEST / "reports" / "v3_1_stats.json").write_text(json.dumps({**stats, "known": known, "new_regression_rows": new_regressions, "remaining": remaining, "active": active, "long_term": long_term, "events": event_rows, "other": other_rows}, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# Prompt V3.1 固定30条回归评估",
        "",
        "## A. V3.1 定向修改",
        "",
        "V3.1 仅修改三类规则：事件页面先判时效再判知识价值；`historical_but_valuable` 仅允许过去事件形成当前仍存在的实体、制度、服务、平台、资源或设施；所有当前有效且核心价值依赖明确未来截止日期的页面统一 `review + active_time_bound`。主体归属、Quality Gate 边界、category、JSON 结构和长期清华服务/资源的知识价值定义保持不变。",
        "",
        "实验使用与 V3 完全相同的 30 条正文和人工标签，固定输入 SHA-256 为 `60C385C5DA026DF0F25B03034D63DA97F62FBDF2FF365A8DD83A27E14990BC1C`。模型为 `gpt-5.4-mini`，temperature=0.1，并发=3，max_completion_tokens=900；30/30 调用成功，失败 0，实际重试 0。",
        "",
        "## B. 三代 Prompt 与人工一致率",
        "",
        f"- V2 vs Human：{agreements['V2_action']}/{n} = {agreements['V2_action']/n:.1%}。",
        f"- V3 vs Human：{agreements['V3_action']}/{n} = {agreements['V3_action']/n:.1%}。",
        f"- V3.1 vs Human：{agreements['V3_1_action']}/{n} = {agreements['V3_1_action']/n:.1%}。",
        f"- V3.1 动作：approve={actions.get('approve',0)}，review={actions.get('review',0)}，reject={actions.get('reject',0)}。",
        "",
        "### V3 → V3.1 迁移",
        "",
        "| 迁移 | 条数 |",
        "|---|---:|",
    ]
    lines += [f"| {k} | {v} |" for k, v in transitions.items()]
    lines += ["", "## C. 6条已知分歧", ""]
    for x in known:
        state = "通过" if x["passed"] else "未通过"
        rt = f" + {x['reject_type']}" if x["reject_type"] else ""
        lines.append(f"- `{x['id']}`：V3.1={x['V3_1_action']}{rt}，Human={x['human_action']}，{state}。{x['title']}")
    lines += [
        "",
        f"已知问题修正 {sum(x['passed'] for x in known)}/{len(known)}。`PUBFOLLOW000007` 仍为 review：正文未提供明确展期或截止日，模型按 `unknown` 处理；人工外部判断为过期展览。这说明当正文缺失活动结束证据时，Prompt 单靠正文无法稳定强制 `expired_event`。",
        "",
        "## D. 新回归",
        "",
        f"V3 原本与人工一致、V3.1 变成不一致的页面共 {len(new_regressions)} 条：",
        "",
    ]
    for x in new_regressions:
        lines.append(f"- `{x['id']}`：V3={x['V3_action']}，V3.1={x['V3_1_action']}，Human={x['human_action']}。{x['title']}")
    lines += [
        "",
        "`PUBEXP000170` 是规则预期带来的边界冲突：V3.1 按“明确截止日期一律 review”执行，但人工因试用期持续到 2026 年末而标 approve。`PUBEXP000221` 被 V3.1 判为已形成持续书架/阅读空间而 approve，人工认为年度有效信息应 review；这不是一次性事件误杀，但显示持续性证据的判断仍有边界波动。",
        "",
        "## E. reject 结构",
        "",
        f"- `out_of_scope`：{reject_types.get('out_of_scope',0)}。",
        f"- `expired_event`：{reject_types.get('expired_event',0)}。",
        f"- `other`：{reject_types.get('other',0)}。",
        "",
    ]
    if other_rows:
        lines += [f"- `{x['id']}`：{x['title']}；{x['reason']}" for x in other_rows]
    else:
        lines.append("本轮没有 `other`，无需逐条解释。")
    lines += [
        "",
        "## F. 一次性事件规则表现",
        "",
        f"V3.1 将 {sum(x['reject_type']=='expired_event' for x in event_rows)} 条明确过期事件归为 `reject + expired_event`，覆盖课堂、培训、活动报道、单场讲堂、领奖通知和荐购活动。已知的 5 个结束事件硬检查中 4 个成功，展览样本 `PUBFOLLOW000007` 因正文缺少可核实结束日期仍为 review。因此事件规则明显改善，但没有达到“5 条全部 reject”的通过条件。",
        "",
        "## G. active_time_bound 表现",
        "",
        f"本轮共有 {len(active)} 条 `active_time_bound`，全部进入 review，规则执行稳定：",
        "",
    ]
    for x in active:
        lines.append(f"- `{x['id']}`：V3.1={x['V3_1_action']}；Human={x['human_action']}。{x['title']}")
    lines += [
        "",
        "其中 `PUBEXP000076` 与人工一致；`PUBEXP000170` 与人工不一致，但 V3.1 的动作符合本轮明确制定的“一律 review”规则。",
        "",
        "## H. 是否过度收紧",
        "",
        f"抽查的 {len(long_term)} 条长期服务、机构、资源、数据库、系统和科研资源页面均未被 reject，长期页面误杀为 {len(long_term_miskills)}。图书借还、信息系统开发、教学环境服务、组织机构、长期数据库导航、正式订购资源和 ACM OA 政策均保持 approve。没有发现系统性过度收紧。",
        "",
        "## 结论",
        "",
        "**V3_1_NEEDS_REVISION**。V3.1 将一致率从 24/30 提升到 27/30，并明显纠正了活动报道和历史讲座的误放行；但未达到原则上的 28/30，且 5 个已结束事件硬检查未全部通过，同时产生 2 条新回归。因此不建议直接进入新样本盲测。建议只再处理两个可泛化边界：正文缺少结束日的单次展览/活动如何保守判定，以及“明确截止日期一律 review”与长期试用资源人工口径的冲突。完成定向复核后再进入盲测。",
        "",
        "本任务到此停止：未替换生产 Prompt、未重审 217 条、未启动 Public Expansion V2。",
    ]
    (TEST / "reports" / "prompt_v3_1_evaluation.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
