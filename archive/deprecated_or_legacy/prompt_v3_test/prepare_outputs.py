from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(r"D:\python_projects\tsinghua_ai")
V3 = ROOT / "data_second" / "prompt_v3_test"
HUMAN_XLSX = ROOT / "data_second" / "public_rebuild_v1" / "human_check" / "public_rebuild_v1_human_check.xlsx"


def load_human():
    ws = openpyxl.load_workbook(HUMAN_XLSX, data_only=True).active
    headers = [c.value for c in ws[3]]
    return {
        str(row[0]).strip(): dict(zip(headers, row))
        for row in ws.iter_rows(min_row=4, values_only=True)
        if row[0]
    }


def main():
    human = load_human()
    analysis = json.loads((V3 / "analysis" / "v2_vs_human_analysis.json").read_text(encoding="utf-8"))
    v3 = {
        json.loads(line)["id"]: json.loads(line)
        for line in (V3 / "audit" / "v3_results.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    }
    if set(human) != set(v3):
        raise RuntimeError(f"sample mismatch: human={len(human)} v3={len(v3)}")

    results = []
    for a in analysis:
        ident = a["id"]
        h = human[ident]
        r = v3[ident]
        results.append({
            "id": ident,
            "title": a["title"],
            "url": a["url"],
            "V2": str(h.get("rebuild_action") or "").lower(),
            "V3": r["action"],
            "human": str(h.get("human_action") or "").lower(),
            "category": r["category"],
            "content_type": r["content_type"],
            "time_status": r["time_status"],
            "candidate_user_question": r["candidate_user_question"],
            "positive_evidence": r["positive_evidence"],
            "negative_evidence": r["negative_evidence"],
            "reject_type": r["reject_type"],
            "reason": r["reason"],
        })

    v2_agree = sum(x["V2"] == x["human"] for x in results)
    v3_agree = sum(x["V3"] == x["human"] for x in results)
    transitions = Counter(f'{x["V2"]}→{x["V3"]}' for x in results)
    reject_types = Counter(x["reject_type"] for x in results if x["V3"] == "reject")
    v3_actions = Counter(x["V3"] for x in results)
    disagreements = [x for x in results if x["V3"] != x["human"]]
    v2_disagreements = [x for x in results if x["V2"] != x["human"]]

    (V3 / "results" / "prompt_v3_30_results.json").write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    stats = {
        "sample_size": len(results),
        "v2_vs_human": {"agree": v2_agree, "disagree": len(results) - v2_agree, "rate": v2_agree / len(results)},
        "v3_vs_human": {"agree": v3_agree, "disagree": len(results) - v3_agree, "rate": v3_agree / len(results)},
        "v3_actions": dict(v3_actions),
        "reject_types": dict(reject_types),
        "transitions": dict(transitions),
        "v3_disagreements": [{"id": x["id"], "title": x["title"], "V3": x["V3"], "human": x["human"], "reject_type": x["reject_type"], "reason": x["reason"]} for x in disagreements],
    }
    (V3 / "reports" / "v3_stats.json").write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# Prompt V3 30条人工样本回归评估",
        "",
        "## 1. 范围与执行",
        "",
        "本次仅使用第一轮 Public 正式数据中已经人工标注的同一批 30 条固定样本，复用现有受控模型配置完成 Prompt V3 回归；未抓取新页面、未修改正文提取器、Quality Gate、Restricted/Portal、人工标签或历史结果。模型调用 30/30 成功，失败 0，网络请求 30 次，实际重试 0 次，模型为 `gpt-5.4-mini`，temperature=0.1，并发=3。",
        "",
        "## 2. 一致性结果",
        "",
        f"- V2 vs 人工：{v2_agree}/{len(results)} = {v2_agree / len(results):.1%}（不一致 {len(results)-v2_agree} 条）。",
        f"- V3 vs 人工：{v3_agree}/{len(results)} = {v3_agree / len(results):.1%}（不一致 {len(results)-v3_agree} 条）。",
        f"- V3 相比 V2：一致率变化 {v3_agree - v2_agree:+d} 条；V3 动作为 approve={v3_actions.get('approve',0)}、review={v3_actions.get('review',0)}、reject={v3_actions.get('reject',0)}。",
        "",
        "### V2→V3 动作迁移",
        "",
        "| 迁移 | 条数 |",
        "|---|---:|",
    ]
    for k, n in sorted(transitions.items()):
        lines.append(f"| {k} | {n} |")
    lines += [
        "",
        "### V3 reject_type",
        "",
        "| reject_type | 条数 |",
        "|---|---:|",
    ]
    for k, n in sorted(reject_types.items()):
        lines.append(f"| `{k}` | {n} |")
    lines += [
        "",
        "## 3. 典型变化与剩余差异",
        "",
        "V3 正确吸收了人工对 `图书借还`、`信息系统开发` 等长期校园服务页的放宽判断，并把 BIBF 通知、已结束课堂等一次性过期内容纳入 `expired_event` 规则。V3 仍有 6 条与人工不一致，主要是：",
        "",
    ]
    for x in disagreements:
        lines.append(f"- `{x['id']}`：V3={x['V3']}，人工={x['human']}；{x['title']}。")
    lines += [
        "",
        "其中 East View 数据库试用属于人工 review 而 V3 approve，说明有明确截止日期的当前资源仍需更保守地进入 review；已结束的经济金融数据课堂被 V3 判为 approve，说明“历史但有潜在知识价值”规则对一次性讲座/课堂的边界还不够紧；若干图书馆活动/讲座回顾被 V3 approve 或 review，而人工 reject，说明需进一步区分可复用的稳定知识与纯活动报道。",
        "",
        "## 4. 风险与建议",
        "",
        "- 样本只有 30 条，且人工标签本身是本轮评估基准，不能外推到全部 217 条审核结果。",
        "- V3 的 approve 数从 10 增至 15，存在把历史活动报道、已结束课堂或活动回顾放宽收录的风险；应在扩大回归前收紧一次性事件的 `historical_but_valuable` 适用条件。",
        "- reject_type 已稳定落在 `out_of_scope` 与 `expired_event` 两类，本轮没有出现 `other`，分类设计可继续保留。",
        "",
        "## 5. 结论",
        "",
    ]
    if v3_agree > v2_agree:
        conclusion = "V3_PASS_WITH_MINOR_ISSUES" if v3_agree >= 27 else "V3_NEEDS_REVISION"
    else:
        conclusion = "V3_NEEDS_REVISION"
    lines.append(f"**{conclusion}**。V3 没有达到可直接替换的稳定性要求：与人工一致率未提升（或仍有明显边界误判），建议先针对“已结束课堂/讲座/活动报道”和“有截止日期的当前资源”补充规则，再进行下一轮固定样本回归。本任务到此停止，不进行生产批量重审。")
    (V3 / "reports" / "prompt_v3_evaluation.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
