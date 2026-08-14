from __future__ import annotations

import json
import re
from pathlib import Path
from collections import Counter

import openpyxl


ROOT = Path(r"D:\python_projects\tsinghua_ai")
OUT = ROOT / "data_second" / "public_rebuild_v1"
V3 = ROOT / "data_second" / "prompt_v3_test"
V3.mkdir(parents=True, exist_ok=True)
for rel in ("analysis", "prompt", "results", "reports", "audit"):
    (V3 / rel).mkdir(parents=True, exist_ok=True)


def load_sheet(path: Path) -> list[dict]:
    ws = openpyxl.load_workbook(path, data_only=True).active
    headers = [ws.cell(3, c).value or "" for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(4, ws.max_row + 1):
        rows.append({headers[c - 1]: ws.cell(r, c).value or "" for c in range(1, ws.max_column + 1)})
    return rows


human_rows = load_sheet(OUT / "human_check" / "public_rebuild_v1_human_check.xlsx")
audit_rows = [json.loads(line) for line in (OUT / "audit" / "audit_results.jsonl").read_text(encoding="utf-8").splitlines() if line.strip()]
quality = [json.loads(line) for line in (OUT / "intermediate" / "base_quality.jsonl").read_text(encoding="utf-8").splitlines() if line.strip()]
follow = [json.loads(line) for line in (OUT / "intermediate" / "follow_quality.jsonl").read_text(encoding="utf-8").splitlines() if line.strip()]
all_meta = {r["id"]: r for r in quality + follow}
audit_by_id = {r["id"]: r for r in audit_rows}


def reason_rule(row: dict, v2: str, human: str, note: str) -> tuple[str, str]:
    text = f"{row.get('title','')} {note}".lower()
    expired = bool(re.search(r"过时|结束|已结束|失效|过去|早已|截止|展览|讲座|活动|通知", text))
    external = bool(re.search(r"其他学校|别的学校|外校|行业|与主题没有任何关系|关系不大", text))
    durable = bool(re.search(r"借阅|机构|服务|信息系统|图书馆|数据库|公共服务|基本信息|规则|管理办法|资源|平台", text))
    if v2 == human:
        return "一致，无需新增规则。", "保持现有 V2 规则。"
    if human == "reject" and expired:
        return "人工认为页面核心价值依赖已结束的一次性活动、通知、展览或讲座，V2 对时效性放得过宽。", "引入 expired_event：已结束且未形成持续知识的单次事件直接 reject。"
    if human == "reject" and external:
        return "人工认为页面主体是其他高校或泛行业内容，和清华知识体系缺乏实质关系，V2 的主题归属边界不够明确。", "先判断页面主体是否属于清华自身知识体系；外校制度、外校政策和泛行业内容归 out_of_scope。"
    if human == "approve" and v2 == "reject" and durable:
        return "人工认为这是清华机构、公共服务、图书馆规则或信息系统等长期知识，V2 把非办事型知识误杀。", "扩大长期清华知识准入：机构介绍、公共服务、资源、规则、科研体系和校园运行事实可 approve。"
    if human == "review" and v2 == "approve":
        return "人工认为页面有价值但存在明确有效期或当前有效性不确定，V2 直接 approve 过宽。", "对 active_time_bound 或有效性无法确认的高价值内容优先 review；记录截止日期。"
    if human == "review" and v2 == "reject":
        return "人工认为页面可能是持续中的服务或长期专题，但当前有效性/主体边界需要确认，V2 直接 reject 过严。", "对有持续价值但有效期或主体边界不确定的内容使用 review，不把 review 当低价值兜底。"
    if human == "approve" and v2 == "reject":
        return "人工认为页面包含可复用的清华事实、服务或资源入口，不能仅因新闻/宣传形式或缺少完整办事流程而 reject。", "新闻形式不自动 reject；只要包含当前仍成立的清华事实、服务能力、资源名称或机构职责，可 approve。"
    return "人工与 V2 的差异指向主题边界、时效性或知识价值判断，需要显式拆分。", "使用主体归属 → 长期/当前价值 → 时效性 → 形式的顺序判断。"


analysis_rows = []
v3_inputs = []
for h in human_rows:
    ident = h["id"]
    meta = {**all_meta.get(ident, {}), **audit_by_id.get(ident, {})}
    v2 = h.get("rebuild_action", "")
    human_action = h.get("human_action", "")
    note = h.get("human_note", "")
    diff_reason, proposed_rule = reason_rule(meta, v2, human_action, note)
    analysis_rows.append({
        "id": ident, "title": h.get("title", ""), "url": h.get("url", ""), "V2_action": v2,
        "human_action": human_action, "content_type": h.get("content_type", meta.get("content_type", "")),
        "time_status": meta.get("time_status", ""), "human_note": note,
        "disagreement_reason": diff_reason, "proposed_v3_rule": proposed_rule,
    })
    source_file = meta.get("source_file", "")
    content = (OUT / source_file).read_text(encoding="utf-8") if source_file else ""
    v3_inputs.append({
        "id": ident, "parent_list_id": meta.get("parent_list_id", ""), "title": h.get("title", ""),
        "url": h.get("url", ""), "source_domain": meta.get("source_domain", ""),
        "source_file": source_file, "crawl_time": meta.get("crawl_time", ""),
        "extraction_method": meta.get("extraction_method", ""), "content_quality_class": meta.get("content_quality_class", ""),
        "content_type_v2": h.get("content_type", meta.get("content_type", "")),
        "category_v2": h.get("category", meta.get("category", "")), "time_status_v2": meta.get("time_status", ""),
        "candidate_user_question_v2": meta.get("candidate_user_question", ""),
        "positive_evidence_v2": meta.get("positive_evidence", ""), "negative_evidence_v2": meta.get("negative_evidence", ""),
        "v2_reason": meta.get("reason", ""), "content": content,
    })

(V3 / "analysis" / "v2_vs_human_analysis.json").write_text(json.dumps(analysis_rows, ensure_ascii=False, indent=2), encoding="utf-8")
(V3 / "audit" / "v3_inputs.json").write_text(json.dumps(v3_inputs, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({"human_rows": len(human_rows), "disagreements": sum(r["V2_action"] != r["human_action"] for r in analysis_rows), "v2": dict(Counter(r["V2_action"] for r in analysis_rows)), "human": dict(Counter(r["human_action"] for r in analysis_rows))}, ensure_ascii=False))
