from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from .common import dedupe_path, ensure_output_path
from .models import FilePlan


def _display_width(value: Any) -> int:
    text = str(value) if value is not None else ""
    return sum(2 if "㐀" <= char <= "鿿" else 1 for char in text)


def _style_sheet(sheet: Any) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="D9DEE7")
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10.5)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.value is not None:
                cell.border = Border(bottom=thin)
    if sheet.max_row:
        for cell in sheet[1]:
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="6F2C91")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        length = max((_display_width(cell.value) for cell in column), default=0)
        sheet.column_dimensions[letter].width = min(max(length + 3, 12), 42)
    for row in range(1, sheet.max_row + 1):
        estimated_lines = 1
        for cell in sheet[row]:
            width = max(1, int(sheet.column_dimensions[cell.column_letter].width or 12))
            estimated_lines = max(estimated_lines, (_display_width(cell.value) + width - 1) // width)
        sheet.row_dimensions[row].height = max(24, 18 * estimated_lines)


def create_xlsx(plan: FilePlan, output_path: str | Path | None = None, template_path: str | Path | None = None) -> Path:
    from openpyxl import Workbook, load_workbook

    path = dedupe_path(ensure_output_path(plan.title, "xlsx", output_path))
    workbook = load_workbook(template_path) if template_path else Workbook()
    if not template_path:
        workbook.remove(workbook.active)
    sheets = plan.workbook_sheets
    if not sheets:
        sheets = []
        for section in plan.sections:
            rows = section.table or [["项目", "内容"]]
            if not section.table:
                for paragraph in section.paragraphs:
                    rows.append([section.heading, paragraph])
                for bullet in section.bullets:
                    rows.append([section.heading, bullet])
            sheets.append({"name": section.heading[:31] or "内容", "rows": rows})
    if not sheets:
        sheets = [{"name": "内容", "rows": [["项目", "内容"], ["标题", plan.title]]}]
    existing = set(workbook.sheetnames)
    for sheet_spec in sheets:
        name = str(sheet_spec.get("name", "Sheet"))[:31] or "Sheet"
        base, index = name, 2
        while name in existing:
            name = f"{base[:27]}_{index}"
            index += 1
        existing.add(name)
        sheet = workbook.create_sheet(name)
        for row in sheet_spec.get("rows", []):
            sheet.append(list(row))
        for address, formula in sheet_spec.get("formulas", {}).items():
            sheet[address] = formula
        _style_sheet(sheet)
    workbook.properties.title = plan.title
    workbook.properties.creator = plan.author or "清问·TsingAsk V2"
    workbook.save(path)
    return path


def modify_xlsx(
    input_path: str | Path,
    *,
    cell_updates: dict[str, Any] | None = None,
    replacements: dict[str, str] | None = None,
    output_path: str | Path | None = None,
) -> tuple[Path, int]:
    from openpyxl import load_workbook

    source = Path(input_path).resolve()
    path = dedupe_path(ensure_output_path(f"{source.stem}_modified", "xlsx", output_path))
    shutil.copy2(source, path)
    workbook = load_workbook(path)
    changed = 0
    for qualified, value in (cell_updates or {}).items():
        if "!" not in qualified:
            raise ValueError(f"cell update must use Sheet!A1 syntax: {qualified}")
        sheet_name, address = qualified.rsplit("!", 1)
        workbook[sheet_name][address] = value
        changed += 1
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or cell.value.startswith("="):
                    continue
                updated = cell.value
                for old, new in (replacements or {}).items():
                    updated = updated.replace(old, new)
                if updated != cell.value:
                    cell.value = updated
                    changed += 1
    workbook.save(path)
    return path, changed


def read_xlsx(path: str | Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        sheets.append({"name": sheet.title, "rows": rows, "max_row": sheet.max_row, "max_column": sheet.max_column})
    workbook.close()
    return {"format": "xlsx", "sheets": sheets, "sheet_count": len(sheets)}
